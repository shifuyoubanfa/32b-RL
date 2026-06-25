"""Step 15（RL 阶段·扩 query）：从"一阶段模型输出"Excel 摄取**全新** query，构成扩充的 RL 训练池。

动机：SFT→RFT 一直在同一批 ~800 query 上薅，RL(尤其 GRPO)有过拟合到该 query 集的风险。
这批 1773 条生产真实问题（带答案+满意度/解决情况评价）我们一条没用过——拿来扩池，既防过拟合，
又能证明"humanness 提升在没训过的新 query 上也成立"(方法泛化，对迁移回 V1 是强证据)。

为什么不需要 think：RL 阶段 think 是模型【自己生成】的优化对象，不是输入。所以新 query 只需
query + 参考资料(当 prompt) + 答案(当准确率锚)——这文件恰好齐了。

可信答案筛选：答案是一阶段模型输出(非人工"可用"硬标注)，只保留 解决情况==解决 或 满意度∈{非常满意,满意}
的行，把答案当准确率锚；排除 不满意/未解决/不确定。

格式对齐：参考资料用 rag_client.build_user_prompt 拼，与 SFT/RFT 训练时【完全一致】的壳
（reward 靠 extract_references 从 【参考问答对】…【问题】 之间切文本做照抄检测）。

去重：与验收集 SFT_EVAL 严格去重(防泄漏)；与现有 SFT_TRAIN 去重(避免重复)。
默认把 SFT_TRAIN 的 query 一并并入输出，形成"老池 + 新池"的扩充训练池。

输出 RL_POOL_EXPANDED，每行：{query, user_prompt, answer, messages:[system, user]}（step08/step10 认这个）。
"""

import argparse
import json
import random
import re
import sys
from difflib import SequenceMatcher
from pathlib import Path

sys.path.append(str(Path(__file__).resolve().parents[1]))
from config import (
    NEW_QUERY_XLSX, RL_POOL_EXPANDED, SFT_TRAIN, SFT_EVAL, SYSTEM_PROMPT,
)
from pipeline.logger import get_logger
from pipeline.rag_client import build_user_prompt

log = get_logger("step15_newq")

# 可信答案：满意度/解决情况里这些算"答案可当准确率锚"
GOOD_SAT = {"非常满意", "满意"}
GOOD_RES = {"解决"}
BAD = {"不满意", "未解决", "不确定"}


def norm_q(q: str) -> str:
    """query 归一化（去空白/标点/大小写）用于跨源去重。"""
    s = re.sub(r"\s+", "", (q or ""))
    s = re.sub(r"[，。？！,.\?!；;：:、\"'《》（）()【】\[\]]", "", s)
    return s.lower()


def is_near_dup(qn: str, key_list: list[str], thresh: float = 0.85) -> bool:
    """qn 与 key_list 里任一是否近似重复(改写/同义/语序)。
    用于防验收集泄漏：norm_q 精确匹配挡不住"个税怎么算 vs 如何计算个人所得税"，加模糊一层。"""
    if not qn:
        return False
    sm = SequenceMatcher()
    sm.set_seq2(qn)                       # 固定 b，复用 b2j，快
    for k in key_list:
        if not k:
            continue
        sm.set_seq1(k)
        if sm.real_quick_ratio() < thresh or sm.quick_ratio() < thresh:
            continue                      # 上界都不够，跳过昂贵的 ratio()
        if sm.ratio() >= thresh:
            return True
    return False


def parse_refs(raw: str) -> list[dict]:
    """把 参考知识TOP5 的 `[qa][id：id]【标题：内容】[qa]…` 解析成 [{title, content}]。
    解析失败兜底为单条(整段当 content)，保证参考块非空、reward 照抄检测仍有料。"""
    items = []
    if not raw or not str(raw).strip():
        return items
    raw = str(raw)
    chunks = re.split(r"\[qa\]", raw)
    for ch in chunks:
        ch = ch.strip()
        if not ch:
            continue
        m = re.search(r"【(.*)】", ch, re.S)   # 贪婪取 标题：内容
        body = m.group(1) if m else ch
        if "：" in body:
            title, content = body.split("：", 1)
        elif ":" in body:
            title, content = body.split(":", 1)
        else:
            title, content = "", body
        title, content = title.strip(), content.strip()
        if content:
            items.append({"title": title, "content": content})
    if not items:
        items = [{"title": "", "content": raw.strip()}]
    return items


def load_existing_queries(path: str) -> list[dict]:
    out = []
    p = Path(path)
    if not p.exists():
        return out
    with p.open("r", encoding="utf-8") as f:
        for l in f:
            if l.strip():
                try:
                    out.append(json.loads(l))
                except Exception:
                    continue
    return out


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", default=NEW_QUERY_XLSX)
    parser.add_argument("--out", default=RL_POOL_EXPANDED)
    parser.add_argument("--sheet", default="Sheet0")
    parser.add_argument("--top_k", type=int, default=5)
    parser.add_argument("--max_new", type=int, default=0, help="最多取 N 条新 query（0=全部合格）")
    parser.add_argument("--no_existing", action="store_true",
                        help="不并入 SFT_TRAIN，只输出纯新 query（默认并入形成扩充池）")
    args = parser.parse_args()

    import openpyxl
    if not Path(args.xlsx).exists():
        log.error("找不到 Excel: %s", args.xlsx)
        sys.exit(1)

    # 验收集 query（严格排除，防泄漏）+ 现有训练 query（去重）
    eval_recs = load_existing_queries(SFT_EVAL)
    train_recs = load_existing_queries(SFT_TRAIN)
    eval_keys = {norm_q(r.get("query")) for r in eval_recs}
    eval_keys_list = [k for k in eval_keys if k]      # 供模糊查重
    train_keys = {norm_q(r.get("query")) for r in train_recs}
    log.info("验收集 %d 条(精确+模糊排除，防泄漏)，现有训练 %d 条", len(eval_recs), len(train_recs))

    wb = openpyxl.load_workbook(args.xlsx, read_only=True)
    ws = wb[args.sheet] if args.sheet in wb.sheetnames else wb.worksheets[0]
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    ci = {h: i for i, h in enumerate(hdr)}
    need = ["模型总结问题", "用户问题", "参考知识TOP5", "答案", "满意度评价", "解决情况评价"]
    for col in need:
        if col not in ci:
            log.error("Excel 缺列: %s（实际列: %s）", col, hdr)
            sys.exit(1)

    seen = set()                 # 本批内去重
    n_seen = drop_qual = drop_dup = drop_empty = drop_eval = 0
    new_recs = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        q = (r[ci["模型总结问题"]] or r[ci["用户问题"]] or "")
        q = str(q).strip()
        if not q:
            continue
        n_seen += 1
        sat = str(r[ci["满意度评价"]] or "").strip()
        res = str(r[ci["解决情况评价"]] or "").strip()
        # 质量门：解决 或 满意/非常满意，且不在 BAD 里
        if sat in BAD or res in BAD or not (res in GOOD_RES or sat in GOOD_SAT):
            drop_qual += 1
            continue
        ans = str(r[ci["答案"]] or "").strip()
        if not ans:
            drop_empty += 1
            continue
        k = norm_q(q)
        if k in eval_keys or is_near_dup(k, eval_keys_list):   # 精确 + 模糊：严防验收集泄漏
            drop_eval += 1
            continue
        if k in train_keys or k in seen:
            drop_dup += 1
            continue
        seen.add(k)
        items = parse_refs(r[ci["参考知识TOP5"]])
        user_prompt = build_user_prompt(q, items, top_k=args.top_k)
        new_recs.append({
            "query": q,
            "user_prompt": user_prompt,
            "answer": ans,
            "gold_source": "stage1",     # 金标准来自一阶段模型(满意度筛过)，非人工硬标注——可溯源
            "messages": [
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": user_prompt},
            ],
        })
        if args.max_new and len(new_recs) >= args.max_new:
            break
    wb.close()

    log.info("Excel 扫描 %d 行：合格新 query=%d（丢弃：质量%d / 空答案%d / 撞验收集%d / 重复%d）",
             n_seen, len(new_recs), drop_qual, drop_empty, drop_eval, drop_dup)

    # 组装扩充池：现有训练(归一成 RL 池格式) + 新 query
    pool = []
    if not args.no_existing:
        for rec in train_recs:
            msgs = rec.get("messages") or []
            if len(msgs) < 2:
                continue
            pool.append({
                "query": rec.get("query"),
                "user_prompt": msgs[1].get("content", ""),
                "answer": (rec.get("answer") or "").strip(),
                "gold_source": "verified",   # 老池金标准=人工"可用"标注
                "messages": msgs[:2],
            })
    pool.extend(new_recs)
    # 确定性打散老/新：DPO rollout 有 max_queries 上限(默认400)，不打散就只采到前面的老 query、采不到新 query。
    random.seed(42)
    random.shuffle(pool)

    with open(args.out, "w", encoding="utf-8") as fout:
        for rec in pool:
            fout.write(json.dumps(rec, ensure_ascii=False) + "\n")

    log.info("扩充 RL 池写入 %d 条（老 %d + 新 %d）-> %s",
             len(pool), len(pool) - len(new_recs), len(new_recs), args.out)
    if len(new_recs) < 50:
        log.warning("新 query 偏少(%d)，可放宽质量门或检查 Excel 列。", len(new_recs))


if __name__ == "__main__":
    main()
